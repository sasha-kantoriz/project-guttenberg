"""Guttenberg2.py.

usage: python3 guttenberg2.py [options]

Project Guttenberg books scrape script:

options:
  -h, --help            show this help message and exit
  -w, --workers         Number of concurrent workers to use (default: 4)

Script will create output folder named as datestamp, and also maintain last processed book index and Excel file with each run spreadsheet


Metadata needed for the script includes:
- Input info of the two source books (orange headers) reference_id from gutenberg, title and author for the two books
- Output info (green headers): ID, title, description



Interior requirements:
-Title page (bundle title centered) with both authors below → format: Author 1 & Author 2.
-Blank page.
-“Featured books” page (centered header). Then centered in the middle of the page two lines:
-<Title 1>; <Author 1> — "Page " + page number where book 1 starts
-<Title 2>; <Author 2> — "Page" + page number where the book 2 starts
- Blank page
- Title page for book 1 (same format as the bundle but with title 1 and author 1)
- Book 1
- Blank page
- Title page for book 2 (same format as the bundle but with title 2 and author 2)
- Book 2

Cover requirements (same format as current script but with bundle title and the 2 authors)
- Title : bundle title from metadata (column G)
- Authors : author 1 "&" author 2 (column C and D)
- Image: generate with the current prompt, adapting it to use the two titles, author names and description
- Description in back cover: bundle description (column I)

Spreadsheet:
- Three output columns from Metadata (ID Title Description) + total number of pages of the combined interior
"""

import os
import re
import fpdf
import logging
import argparse
import pathlib
import requests
import openpyxl
import traceback
import pandas as pd
import concurrent.futures
from PIL import Image
from time import sleep
from datetime import datetime
from tempfile import TemporaryFile
from openai import OpenAI


client = OpenAI()

logger = logging.getLogger("pg-bundles")


def load_current_progress():
    try:
        with open('index', 'r') as f:
            return int(f.read().strip())
    except (FileNotFoundError, ValueError):
        return 0


def dump_current_progress(index):
    with open('index', 'w') as f:
        f.write(str(index))


def fetch_guttenberg_book(index):
    try:
        logger.debug(f"Fetching book {index}")
        book_url = f'https://www.gutenberg.org/ebooks/{index}.txt.utf-8'
        response = requests.get(
            book_url,
            timeout=60,
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; zh-CN) AppleWebKit/533+ (KHTML, like Gecko)'
            }
        )
        #
        if response.status_code != 200:
            logger.error(f"Error fetching book text: {response.status_code}")
            return None
        #
        logger.debug(f"Successfully fetched book {index}")
        #
        book_txt = response.content.decode('utf-8')
        return book_txt
    except:
        return None


def format_contents_with_openai(raw_contents):
    """
    Formats the raw contents section using OpenAI API.

    Args:
        raw_contents (str): The unformatted "Contents" section.

    Returns:
        str: The formatted "Contents" section.
    """

    book_contents = raw_contents
    prompt = f"""
    The following text is a raw Contents section from a book. Please format it into a clean and structured "Contents" section while adhering to these guidelines:
    1. Retain the original Roman numerals for chapter numeration, if present.
    2. Remove all page numbers from the text.
    3. Align chapters to the left, without indentation
    4. Ensure consistent spacing and alignment for all lines.
    5. Preserve the order and structure of the titles as they appear in the raw input.
    6. If the word "Index" appears at the end of the "Contents" section, remove it entirely.
    7. Do not add "" at the beginning and end of the formatted section

    Here is the raw input:
    {raw_contents}

    Please return the formatted Contents section.
    """
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant for formatting a Contents section text from a book."},
                {"role": "user", "content": prompt},
            ],
            temperature=0  # For consistent formatting
        )
        book_contents = response.choices[0].message.content

    except Exception as e:
        logger.error(f"Error formatting contents with OpenAI API: {e}")

    return book_contents  # Return raw contents if API call fails


def parse_raw_book(text):
    book_author = re.search(r"(Author|Editor): (.*)\r\n", text, re.IGNORECASE)
    book_author = book_author.groups()[1] if book_author else ""
    book_author = book_author.strip().replace('\\', '-').replace('/', '-').replace('&', ' and ')
    book_language = re.search(r"Language: (.*)\r\n", text, re.IGNORECASE)
    book_language = book_language.groups()[0] if book_language else ""
    book_translator = re.search(r"Translator: (.*)\r\n", text, re.IGNORECASE)
    book_translator = book_translator.groups()[0] if book_translator else ""
    book_illustrator = re.search(r"Illustrator: (.*)\r\n", text, re.IGNORECASE)
    book_illustrator = book_illustrator.groups()[0] if book_illustrator else ""
    book_title = re.search(r"Title: (.*)\r\n", text)
    book_title = book_title.groups()[0] if book_title else ""
    book_title = book_title.strip().replace('\\', '-').replace('/', '-').replace('&', ' and ')
    book_content_start_index = re.search(r"\*\*\* START OF THE PROJECT GUTENBERG .* \*\*\*", text, re.IGNORECASE)
    book_content_start_index = book_content_start_index.end() if book_content_start_index else 0
    book_content_end_index = re.search(r"\*\*\* END OF THE PROJECT GUTENBERG .* \*\*\*", text, re.IGNORECASE)
    book_content_end_index = book_content_end_index.start() if book_content_end_index else -1
    text = text[book_content_start_index:book_content_end_index]
    # Illustrations supporting text removal
    illustrations_patterns = [
        re.compile(r'\[(\s+)?Cover Illustration](\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[(\s+)?Illustration](\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[(\s+)?Illustration.+?](\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[(\s+)?Ilustracion.+?](\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[(\s+)?Ilustración.+?](\r\n){2}', re.IGNORECASE | re.DOTALL),
    ]
    for _pattern in illustrations_patterns:
        text = re.sub(_pattern, '', text)
    # Proofread text removal
    proofread_patterns = [
        re.compile(r'Produced(.+?)?(\s+)?at(\s+)?(https://|http://)?(www\.)?pgdp\.net(\s+)?(.+?)?(\r\n){3}', re.IGNORECASE | re.DOTALL),
        re.compile(r'Produced(.+?)?(\s+)?by(\s+)?(www\.)?ebooksgratuits\.com(\s+)?(.+?)?(\r\n){3}', re.IGNORECASE | re.DOTALL),
        re.compile(r'(this\s+)?E(-)?(text|book)(\s+)?(is|was)?(\s+)?(produced|prepared)(\s+)?(.+?)?(\r\n){3}', re.IGNORECASE | re.DOTALL),
    ]
    for _pattern in proofread_patterns:
        text = re.sub(_pattern, '', text)
    # Supporting text removal
    produced_by_search = re.search(r'Produced(\s+)?by(\s+)?(.+?)?(\s+)?(.+?)?(\r\n){2}', text[:int(len(text) * 0.05)], re.IGNORECASE | re.DOTALL)
    if produced_by_search:
        text = text.replace(produced_by_search.group(0), '', )
    # Transcriber notes removal
    transcriber_notes_patterns = [
        re.compile(r'(\[)?(\+)?(-{3,}(\+)?)?(\s+)?(\|)?Transcriber(\'s|’s)?(\s+)?Note(s)?(\s+)?(:)?(\+)?(\s+)?(.+?)?(\r\n){3}', re.IGNORECASE | re.DOTALL),
        re.compile(r'Notes de transcription:(\s+)?(:)?(\+)?(\s+)?(.+?)?(\r\n){3}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[Sidenote(s)?(\s+)?:(\s+)?(.+?)?(\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'\[Note(s)?(\s+)?:(\s+)?(.+?)?(\r\n){2}', re.IGNORECASE | re.DOTALL),
    ]
    for _pattern in transcriber_notes_patterns:
        text = re.sub(_pattern, '', text)
    # Removal of project guttenberg marks
    start_end_patterns = [
        re.compile(r'START(\s+)?OF(\s+)?(THE)?(\s+)?PROJECT(\s+)?GUTENBERG.+?(\r\n){2}', re.IGNORECASE | re.DOTALL),
        re.compile(r'END(\s+)?OF(\s+)?(THE)?(\s+)?PROJECT(\s+)?GUTENBERG.+?(\r\n){2}', re.IGNORECASE | re.DOTALL),
    ]
    for _pattern in start_end_patterns:
        text = re.sub(_pattern, '', text)
    #
    text = text.replace('\r\n', '\n')
    # BOOK PUBLISHER NOTES
    book_publisher_notes_start_index, book_publisher_notes_end_index = 0, text[100:int(len(text) * 0.02)].find('\n\n\n\n')
    if book_publisher_notes_end_index != -1:
        book_publisher_notes_end_index += 100
    else:
        book_publisher_notes_end_index = 0
    book_publisher_notes = text[book_publisher_notes_start_index:book_publisher_notes_end_index]
    include_publisher_notes = book_language.lower() not in ['english']
    # BOOK CONTENTS
    contents_search = re.search(
        r"\s+(_)?(table\s+des\s+matières|contenu|liste\s+des\s+matières|contenidos|Índice|Tabla\s+de\s+contenidos|capítulos|list\s+of\s+contents|table\s+of\s+contents|content|contents|contents of volume|contents of volume [IVX]{1,3}|contents of vol|contents of vol(\.)?(\s+[IVX]{1,3})?|chapters|file numbers)(:)?(\.)?(_)?(\n){2,}",
        text[:int(len(text) * 0.15)], re.IGNORECASE | re.DOTALL
    )
    if contents_search and not re.search(r"(content|contents|chapters|file numbers)(:)?(\.)?(\n)+(\s)*of", text[:contents_search.start() + 100], re.IGNORECASE):
        contents_start_index = contents_search.start()
        contents_end_index = contents_start_index + len(contents_search.group()) + 5 + text[contents_start_index + len(contents_search.group()) + 5:].find('\n\n\n\n')
    else:
        contents_end_index = contents_start_index = 0
    book_contents = text[contents_start_index:contents_end_index]
    # Book preface
    preface_search = re.search(
        r'(preface|foreword|prefatory note|préface|vorwort|prólogo|prefacio|prefazione)(\.)?(\n){2}',
        text[:int(len(text) * 0.15)], re.IGNORECASE)
    if preface_search:
        preface_start_index = preface_search.start()
        preface_end_index = preface_start_index + len(preface_search.group()) + 10 + text[
            preface_start_index + len(preface_search.group()) + 10:].find('\n\n\n\n')
        book_preface = text[preface_start_index:preface_end_index]
    else:
        preface_end_index = 0
        book_preface = ""
    # check if sections are separated by 3 newlines
    if book_publisher_notes_end_index == contents_end_index == preface_end_index:
        # BOOK PUBLISHER NOTES
        book_publisher_notes_start_index, book_publisher_notes_end_index = 0, text[
            100:int(len(text) * 0.02)].rfind('\n\n\n')
        if book_publisher_notes_end_index != -1:
            book_publisher_notes_end_index += 100
        else:
            book_publisher_notes_end_index = 0
        book_publisher_notes = text[book_publisher_notes_start_index:book_publisher_notes_end_index]
        # BOOK CONTENTS
        contents_search = re.search(
            r"\s+(_)?(table\s+des\s+matières|contenu|liste\s+des\s+matières|contenidos|Índice|Tabla\s+de\s+contenidos|capítulos|list\s+of\s+contents|table\s+of\s+contents|content|contents|contents of volume|contents of volume [IVX]{1,3}|contents of vol|contents of vol(\.)?(\s+[IVX]{1,3})?|chapters|file numbers)(:)?(\.)?(_)?(\n){2,}",
            text[:int(len(text) * 0.15)], re.IGNORECASE | re.DOTALL
        )
        if contents_search and not re.search(r"(content|contents|chapters|file numbers)(:)?(\.)?(\n)+(\s)*of", text[:contents_search.start() + 100], re.IGNORECASE):
            contents_start_index = contents_search.start()
            contents_end_index = contents_start_index + len(contents_search.group()) + 5 + text[contents_start_index + len(contents_search.group()) + 5:].find('\n\n\n')
        else:
            contents_end_index = contents_start_index = 0
        book_contents = text[contents_start_index:contents_end_index]
        preface_search = re.search(r'(_)?(preface|foreword|prefatory note)(\.)?(_)?(\n){2}', text[:int(len(text) * 0.15)], re.IGNORECASE)
        if preface_search:
            preface_start_index = preface_search.start()
            preface_end_index = preface_start_index + len(preface_search.group()) + 10 + text[preface_start_index + len(preface_search.group()) + 10:].find('\n\n\n')
            book_preface = text[preface_start_index:preface_end_index]
        else:
            preface_end_index = 0
            book_preface = ""
    # BOOK INDEX
    appendix_search = re.search(r'(_)?(Index|Index\s+to\s+Letters)(\.)?(:)?(_)?(\n){2}', text[int(len(text) * 0.8):], re.IGNORECASE)
    if appendix_search:
        appendix_start_index = int(len(text) * 0.8) + appendix_search.start()
        # appendix_end_index = appendix_start_index + len(appendix_search.group()) + 10 + text[appendix_start_index + len(appendix_search.group()) + 10:].find('\n\n\n\n')
        # book_appendix = text[appendix_start_index:appendix_end_index]
    else:
        appendix_start_index = len(text)
        # book_appendix = ""
    # Clean book body
    text = text[max(book_publisher_notes_end_index, contents_end_index, preface_end_index):appendix_start_index]
    # Illustrations list
    illustration_list_search = re.search(
        r'(LIST OF ILLUSTRATIONS|List [Oo]f [iI]llustrations|ILLUSTRATIONS OF VOLUME|Illustrations [Oo]f [Vv]olume|ILLUSTRATIONS TO VOLUME|Illustrations [Tt]o [Vv]olume|ILLUSTRATIONS OF VOL|Illustrations [Oo]f [Vv]ol|Illustrations [Tt]o [Vv]ol|ILLUSTRATIONS|Illustrations)(\.)?',
        text[:int(len(text) * 0.15)]
    )
    if illustration_list_search:
        illustrations_start_index = illustration_list_search.start()
        illustrations_end_index = illustrations_start_index + text[illustrations_start_index:].find('\n\n\n\n')
        text = text[illustrations_end_index:]
    # Plates list
    plates_list_search = re.search(
        r'(LIST OF PLATES|List [Oo]f [pP]lates|PLATES OF VOLUME|Plates [Oo]f [Vv]olume)(\.)?',
        text[:int(len(text) * 0.15)]
    )
    if plates_list_search:
        plates_start_index = plates_list_search.start()
        plates_end_index = plates_start_index + text[plates_start_index:].find('\n\n\n\n')
        text = text[plates_end_index:]
    #
    if book_contents and book_contents in book_publisher_notes:
        book_publisher_notes = ""
    book_publisher_notes = book_publisher_notes.replace('\n\n\n\n', '\n\n').replace('_', '').replace('  ', ' ').replace('--', '-').replace('\n\n', '_____').replace('\n', ' ').replace('_____', '\n\n')
    #
    book_contents_header_search = re.search(
        r"(_)?(table\s+des\s+matières|contenu|liste\s+des\s+matières|contenidos|Índice|Tabla\s+de\s+contenidos|capítulos|list\s+of\s+contents|table\s+of\s+contents|contents|content|contents of volume|contents of volume [IVX]{1,3}|contents of vol|contents of vol(\.)?(\s+[IVX]{1,3})?|chapters|file numbers)(:)?(\.)?(_)?(\n{1,})?",
        book_contents,
        flags=re.DOTALL | re.IGNORECASE
    )
    book_contents_header = book_contents_header_search.group() if book_contents_header_search else ''
    book_contents = re.sub(r'page(s)?(\n)?', '', book_contents, flags=re.IGNORECASE)
    book_contents = book_contents.replace(book_contents_header, '').replace('\n\n\n', '\n').replace('\n\n', '\n')
    book_contents_cleaned = ""
    for book_contents_line in book_contents.split('\n'):
        if book_contents_line and not re.search(r'^((\s+)?chapter|part|volume)', book_contents_line, re.IGNORECASE):
            book_contents_cleaned += re.sub(
                r'([IVX]+|\d+)?(\.)?(\s+)?(.+?)(,)?\s+(\d+|[ivx]+(\.)?)$', r'\1\2\3 \4',
                book_contents_line, flags=re.IGNORECASE | re.DOTALL
            ) + '\n'
        elif book_contents_line:
            book_contents_cleaned += book_contents_line + '\n'
    book_contents = book_contents_header + book_contents_cleaned.replace('_', '').replace('  ', ' ').replace('--', '-')
    book_preface = book_preface.replace('\n\n\n\n', '\n\n').replace('_', '').replace('  ', ' ').replace('--', '-').replace('\n\n', '_____').replace('\n', ' ').replace('_____', '\n\n')
    text = text.replace('\n\n\n\n', '\n\n').replace('_', '').replace('  ', ' ').replace('--', '-').replace('\n\n', '_____').replace('\n', ' ').replace('_____', '\n\n')
    return {
        "Title": book_title,
        "Author": book_author,
        "Language": book_language,
        "Translator": book_translator,
        "Illustrator": book_illustrator,
        "Publisher Notes": book_publisher_notes,
        "Contents": format_contents_with_openai(book_contents),
        "Preface": book_preface,
        "Text": text
    }


class PDF(fpdf.FPDF):
    def footer(self):
        if self.page_no() > 4:
            # Go to 1.5 cm from bottom
            self.set_y(-15)
            self.add_font("dejavu-sans", style="", fname="assets/DejaVuSans.ttf")
            self.set_font("dejavu-sans", size=8)
            # Print centered page number
            self.cell(0, 10, f"{self.page_no()}", 0, 0, 'C')


def write_book_pdf(pdf, title, author, language, text, notes, contents, preface):
    ## Title
    pdf.add_page()
    pdf.set_font("dejavu-sans", size=24)
    title_text = f"{title}\n\n{author}"
    lines_num = len(pdf.multi_cell(w=0, align='C', padding=(0, 8), text=title_text, dry_run=True, output="LINES"))
    if lines_num >= 3:
        padding_top = (228.6 - 24 * (lines_num - 1)) / 2
    else:
        padding_top = (228.6 - 24 * lines_num) / 2
    pdf.multi_cell(w=0, align='C', padding=(padding_top, 8, 0), text=title_text)
    include_publisher_notes = language.lower() not in ['english']
    ## Publisher notes
    if notes and include_publisher_notes:
        pdf.add_page()
        pdf.set_font("dejavu-sans", size=9)
        pdf.multi_cell(w=0, h=4, align='J', padding=8, text=notes)
    ## Contents
    if contents:
        pdf.add_page()
        pdf.set_font("dejavu-sans", size=9)
        pdf.multi_cell(w=0, h=4.4, align='J', padding=8, text=contents)
    ## Preface
    if preface:
        pdf.add_page()
        pdf.set_font("dejavu-sans", size=9)
        pdf.multi_cell(w=0, h=4.4, align='J', padding=8, text=preface)
    ## Text
    pdf.add_page()
    pdf.set_font("dejavu-sans", size=9)
    pdf.multi_cell(w=0, h=4.4, align='J', padding=8, text=text)
    return pdf.page_no()


def generate_bundle_interior_pdf(folder, bundle_id, bundle_title, language_1, language_2, title_1, title_2, author_1, author_2, notes_1, notes_2, contents_1, contents_2, preface_1, preface_2, text_1, text_2):
    """
    -Title page (bundle title centered) with both authors below → format: Author 1 & Author 2.
    -Blank page.
    -“Featured books” page (centered header). Then centered in the middle of the page two lines:
    -<Title 1>; <Author 1> — "Page " + page number where book 1 starts
    -<Title 2>; <Author 2> — "Page" + page number where the book 2 starts
    - Blank page
    - Title page for book 1 (same format as the bundle but with title 1 and author 1)
    - Book 1
    - Blank page
    - Title page for book 2 (same format as the bundle but with title 2 and author 2)
    - Book 2
    """
    interior_pdf_fname = f"{folder}/interior/{bundle_id}_paperback_interior.pdf"

    pdf = PDF(format=(152.4, 228.6))
    pdf.add_font("dejavu-sans", style="", fname="assets/DejaVuSans.ttf")

    # Title page
    pdf.add_page()
    pdf.set_font("dejavu-sans", size=24)
    bundle_title_text = f"{bundle_title}\n\nBy\n\n{author_1} & {author_2}"
    lines_num = len(pdf.multi_cell(w=0, align='C', padding=(0, 8), text=bundle_title_text, dry_run=True, output="LINES"))
    if lines_num >= 3:
        padding_top = (228.6 - 24 * (lines_num - 1)) / 2
    else:
        padding_top = (228.6 - 24 * lines_num) / 2
    pdf.multi_cell(w=0, align='C', padding=(padding_top, 8, 0), text=bundle_title_text)

    # Blank page
    pdf.add_page()

    # Featured books page
    pdf.add_page()
    pdf.set_font("dejavu-sans", size=12)
    with TemporaryFile() as book_1_tmp:
        book_1_tmp_pdf = PDF(format=(152.4, 228.6))
        book_1_tmp_pdf.add_font("dejavu-sans", style="", fname="assets/DejaVuSans.ttf")
        featured_text = f"Featured books:\n\n\n\n{title_1}; {author_1} — Page 4\n\n{title_2}; {author_2} — Page {write_book_pdf(book_1_tmp_pdf, title_1, author_1, language_1, text_1, notes_1, contents_1, preface_1) + 6}"
        book_1_tmp_pdf.output(book_1_tmp)
    lines_num = len(pdf.multi_cell(w=0, align='C', padding=(0, 8), text=featured_text, dry_run=True, output="LINES"))
    if lines_num >= 3:
        padding_top = (228.6 - 24 * (lines_num - 1)) / 2
    else:
        padding_top = (228.6 - 24 * lines_num) / 2
    pdf.multi_cell(w=0, align='C', padding=(padding_top, 8, 0), text=featured_text)

    # Blank page
    pdf.add_page()

    # Book # 1
    write_book_pdf(pdf, title_1, author_1, language_1, text_1, notes_1, contents_1, preface_1)

    # Blank page
    pdf.add_page()

    # Book # 2
    write_book_pdf(pdf, title_2, author_2, language_2, text_2, notes_2, contents_2, preface_2)

    #
    pdf.output(interior_pdf_fname)
    pages = pdf.page_no()
    return pages


def generate_bundle_cover_pdf(folder, bundle_id, title, author, description, interior_pages):
    cover_pdf_fname, dalle_cover_img_png = (
        f"{folder}/cover/{bundle_id}_paperback_cover.pdf",
        f"{folder}/images/{bundle_id}_paperback_cover.png"
    )
    # Full cover
    cover_width, cover_height = 152.4 * 2 + interior_pages * 0.05720 + 3.175 * 2, 234.95
    pdf = fpdf.FPDF(format=(cover_width, cover_height))
    pdf.add_font('dejavu-sans', style="", fname="assets/DejaVuSans.ttf")
    pdf.add_page()
    pdf.set_fill_color(r=250,g=249,b=222)
    pdf.rect(h=pdf.h, w=pdf.w, x=0, y=0, style="DF")
    cols = pdf.text_columns(ncols=2, gutter=interior_pages*0.05720 + 1.588*2, l_margin=6.35, r_margin=6.35)
    #
    description_p = cols.paragraph(text_align='L')
    pdf.set_font('dejavu-sans', size=12)
    description_lines = pdf.multi_cell(w=152.4, align='L', padding=(0, 11.175), text=description, dry_run=True, output="LINES")
    description_p.write('\n'.join(description_lines))
    cols.end_paragraph()
    #
    cols.new_column()
    #
    title_p = cols.paragraph(text_align='C')
    pdf.set_font('dejavu-sans', size=24)
    title_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=f"\n\n{title}", dry_run=True, output="HEIGHT")
    title_p.write(f"\n\n{title}")
    cols.end_paragraph()
    #
    separator_text = "\n* * *"
    separator_p = cols.paragraph(text_align='C')
    pdf.set_font('dejavu-sans', size=16)
    separator_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=separator_text, dry_run=True, output="HEIGHT")
    separator_p.write(separator_text)
    cols.end_paragraph()
    #
    author_p = cols.paragraph(text_align='C')
    pdf.set_font('dejavu-sans', size=16)
    author_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=f"\n{author}\n", dry_run=True, output="HEIGHT")
    author_p.write(f"\n{author}")
    cols.end_paragraph()
    #
    include_cover_img = (title_h + separator_h + author_h + 8) < 234.95 - ((234.95 - 40) / 2 + 10)
    if include_cover_img:
        try:
            prompt = f"""Generate an image to be featured in a book cover. 
            Exclude any depictions of books, book covers or written text on the output image. 
            Meeting the criteria mentioned before, the image needs to be based on the following description: {description}
            """
            img_url = client.images.generate(model='dall-e-3', prompt=prompt, n=1, quality="standard").data[0].url
            response = requests.get(img_url)
            with open(dalle_cover_img_png, 'wb') as img:
                img.write(response.content)
            pdf.image(dalle_cover_img_png, x=(152.4 + interior_pages * 0.05720 + 3.175) + (152.4 - 100 - 6.35) / 2 + 5, y=(234.95 - 40) / 2, w=100, h=100)
        except:
            pass
    #
    cols.render()
    pdf.output(cover_pdf_fname)


def generate_bundle_pdfs(folder, bundle_id, index_1, index_2, title_1, title_2, bundle_title, author_1, author_2, description):
    book_1, book_2 = fetch_guttenberg_book(index_1), fetch_guttenberg_book(index_2)
    if not book_1 or not book_2:
        raise Exception(f"Failed to fetch one or both books for bundle {bundle_id} (books {index_1}, {index_2})")

    book_1_data, book_2_data = parse_raw_book(book_1), parse_raw_book(book_2)
    interior_pages = generate_bundle_interior_pdf(
        folder,
        bundle_id,
        bundle_title,
        book_1_data['Language'],
        book_2_data['Language'],
        title_1,
        title_2,
        author_1,
        author_2,
        book_1_data['Publisher Notes'],
        book_2_data['Publisher Notes'],
        book_1_data['Contents'],
        book_2_data['Contents'],
        book_1_data['Preface'],
        book_2_data['Preface'],
        book_1_data['Text'],
        book_2_data['Text']
    )
    generate_bundle_cover_pdf(folder, bundle_id, bundle_title, f"{author_1} & {author_2}", description, interior_pages)
    return interior_pages

def process_bundle(folder, row):
    """Processes a single book bundle row from the metadata."""
    try:
        logger.info(f"Processing bundle ID: {row['ID']}")
        interior_pages = generate_bundle_pdfs(
            folder,
            row["ID"],
            row["Reference_id_1"],
            row["Reference_id_2"],
            row["Title_1"],
            row["Title_2"],
            row["Title"],
            row["Author_1"],
            row["Author_2"],
            row["Description"]
        )
        logger.info(f"Successfully generated bundle for ID: {row['ID']}")
        return {
            "ID": row["ID"],
            "Title": row["Title"],
            "Description": row["Description"],
            "Pages": interior_pages,
            "Error": None
        }
    except Exception as e:
        logger.error(f"Error processing bundle ID {row['ID']}: {e}")
        logger.error(traceback.format_exc())
        return {
            "ID": row["ID"],
            "Title": row["Title"],
            "Description": row["Description"],
            "Pages": 0,
            "Error": str(e)
        }

def main(folder, num_workers):
    start_index = load_current_progress()
    try:
        all_metadata_rows = pd.read_excel("Bundles_Metadata.xlsx", header=0).to_dict(orient="records")
    except FileNotFoundError:
        logger.error("Metadata file 'Bundles_Metadata.xlsx' not found.")
        return

    metadata_to_process = all_metadata_rows[start_index:]

    if not metadata_to_process:
        logger.info("All bundles have been processed.")
        return

    try:
        wb = openpyxl.load_workbook('Project Guttenberg Bundles.xlsx')
        ws = wb["Sheet"]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws.append(
            [
                "Bundle ID",
                "Title",
                "Description",
                "Pages num"
            ]
        )

    processed_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
        future_to_row = {executor.submit(process_bundle, folder, row): row for row in metadata_to_process}

        for future in concurrent.futures.as_completed(future_to_row):
            result = future.result()
            if result:
                if not result.get("Error"):
                    ws.append(
                        [
                            result["ID"],
                            result["Title"],
                            result["Description"],
                            result["Pages"]
                        ]
                    )
                else:
                    ws.append(
                        [
                            result["ID"],
                            result["Title"],
                            f"ERROR: {result['Error']}",
                            0
                        ]
                    )

                processed_count += 1
                # Save progress intermittently
                if processed_count % num_workers == 0:
                    current_progress = start_index + processed_count
                    logger.info(f"Saving progress. Bundles processed in this run: {processed_count}. Total progress: {current_progress}")
                    wb.save('Project Guttenberg Bundles.xlsx')
                    dump_current_progress(current_progress)

    # Final save and progress update
    final_progress = start_index + processed_count
    wb.save('Project Guttenberg Bundles.xlsx')
    dump_current_progress(final_progress)
    logger.info(f"Finished processing. Total bundles processed in this run: {processed_count}. Final progress: {final_progress}")


def parse_args():
    # parse command line arguments
    parser = argparse.ArgumentParser(
        prog='guttenberg-bundles.py',
        usage='python3 %(prog)s [options]',
        description='Project Guttenberg books scrape script:',
        epilog="Script will create output folder named as datestamp, and also maintain last processed bundle index and Excel spreadsheet"
    )
    parser.add_argument('-w', '--workers', type=int, default=4, help='Number of concurrent workers')
    return parser.parse_args()


if __name__ == '__main__':
    # create output folders
    run_folder = datetime.now().strftime('%Y-%B')
    for subdir in ["interior", "images", "cover"]:
        pathlib.Path(f"{run_folder}/{subdir}").mkdir(parents=True, exist_ok=True)

    # Setup logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.getLogger('fpdf').setLevel(logging.ERROR)
    logging.getLogger('fontTools.subset').setLevel(logging.ERROR)
    # Suppress logs from fpdf.svg (for SVG-related warnings)
    logging.getLogger('fpdf.svg').propagate = False

    args = parse_args()
    main(run_folder, args.workers)
